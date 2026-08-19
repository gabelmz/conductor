"""Data Insights — turn any tabular file into a profile, charts, and an AI narrative.

Drop in an Excel / CSV / TSV / JSON / NDJSON file and Conductor auto-profiles it
independently of the product model: column types, missingness, numeric stats,
top categorical values, and a set of ready-to-render charts (bars, histogram,
time series, dimension × measure). An optional AI pass writes a plain-English
executive summary (hosted provider; graceful when no key is configured).

The parsed table is stored in the `insights_datasets` table so it survives
restarts and can be re-opened, exported to CSV, or summarised later.

Router prefix: /api/insights
"""
from __future__ import annotations

import csv
import io
import json
import math
import re
import statistics
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path

# Make sibling backend modules importable whether this file is loaded as
# `backend.insights` (repo root) or `insights` (backend dir on sys.path, uvicorn).
sys.path.insert(0, str(Path(__file__).resolve().parent))

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import Response

import storage

router = APIRouter(prefix="/api/insights", tags=["insights"])

MAX_UPLOAD_BYTES = 50 * 1024 * 1024
MAX_ROWS_STORE = 25000
MAX_COLS_STORE = 300
MAX_PREVIEW = 100
MAX_CHART_ITEMS = 15

SUPPORTED_EXTS = (".xlsx", ".xlsm", ".csv", ".tsv", ".json", ".ndjson", ".jsonl")

# Values treated as "missing" during profiling (raw rows are kept untouched).
MISSING = {"", "none", "n/a", "na", "nan", "null", "-", "--", "n/a", "nil"}

DATE_FORMATS = (
    "%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d",
    "%m/%d/%Y", "%m/%d/%y", "%m/%d/%Y %H:%M", "%d-%m-%Y", "%b %d %Y",
    "%B %d %Y", "%Y-%m-%dT%H:%M:%S",
)

# Numeric columns whose names hint at "money / quantity" are preferred measures.
MONEY_HINTS = (
    "amount", "price", "cost", "total", "revenue", "sales", "fee", "charge",
    "rate", "spend", "value", "paid",
)
QTY_HINTS = (
    "weight", "qty", "quantity", "count", "units", "unit count", "volume",
)

# Dimension columns that make good chart groupings.
DIMENSION_HINTS = (
    "carrier", "status", "market", "state", "city", "country", "category",
    "type", "brand", "service", "level", "class", "mode", "method", "region",
    "commodity", "warehouse",
)

# Column-name tokens that mark an identifier / key (never numeric, never a measure).
ID_TOKENS = {
    "id", "ids", "bol", "pro", "tracking", "number", "num", "zip", "postal",
    "phone", "account", "sku", "asin", "upc", "ean", "gtin", "reference", "ref",
    "serial", "part", "isbn",
}

# Column-name tokens that hint the column is an address / free-text, not a dimension.
LOW_VALUE_TOKENS = {"address", "line1", "line2", "contact", "name", "notes", "description"}

# Date columns preferred as a time axis (ship/created/order beat estimates/pickup).
DATE_HINTS = [
    ("ship", 5), ("created", 4), ("order", 4), ("date", 1), ("deliver", 1),
    ("actual", 1), ("estimated", -3), ("pickup", -1),
]


# ---------------------------------------------------------------------------
# schema
# ---------------------------------------------------------------------------
def init_insights_db() -> None:
    conn = storage._conn()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS insights_datasets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            filename TEXT DEFAULT '',
            columns TEXT DEFAULT '[]',
            profile TEXT DEFAULT '{}',
            rows_json TEXT DEFAULT '[]',
            row_count INTEGER DEFAULT 0,
            col_count INTEGER DEFAULT 0,
            truncated INTEGER DEFAULT 0,
            created_at TEXT NOT NULL
        )
        """
    )
    conn.commit()


# ---------------------------------------------------------------------------
# value helpers
# ---------------------------------------------------------------------------
def _is_missing(v) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and math.isnan(v):
        return True
    s = str(v).strip()
    return s.lower() in MISSING


def _cell_value(v):
    """Normalise a raw cell to a JSON-safe scalar."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    return str(v).strip()


_NUM_UNIT_RE = re.compile(
    r"\s*(lb|lbs|kg|g|oz|in|inch|inches|cm|mm|ft|feet|m|ml|l|gal|hr|hrs|hours|days|pct|%)?\s*$",
    re.IGNORECASE,
)


def _try_number(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    s = re.sub(r"[$\u00a3\u20ac]", "", s)
    s = s.replace(",", "")
    s = _NUM_UNIT_RE.sub("", s).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _try_date(v):
    if isinstance(v, (datetime, date)):
        return True
    s = str(v).strip()
    if not s:
        return False
    for f in DATE_FORMATS:
        try:
            datetime.strptime(s, f)
            return True
        except ValueError:
            continue
    return False


def _parse_date(v):
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    s = str(v).strip()
    for f in DATE_FORMATS:
        try:
            return datetime.strptime(s, f)
        except ValueError:
            continue
    return None


def _short(v, limit=60) -> str:
    s = str(v if v is not None else "")
    return s if len(s) <= limit else s[: limit - 1] + "…"


# ---------------------------------------------------------------------------
# parsing
# ---------------------------------------------------------------------------
def _clean_header(v, used: set[str]) -> str:
    name = str(v).strip() if v is not None else ""
    if not name:
        name = f"Column {len(used) + 1}"
    base, i = name, 2
    while name in used:
        name = f"{base} ({i})"
        i += 1
    used.add(name)
    return name


def _rows_to_dicts(header: list[str], raw_rows: list[list]) -> list[dict]:
    out = []
    for row in raw_rows:
        if not any(not _is_missing(v) for v in row):
            continue
        d = {}
        for j, name in enumerate(header):
            v = row[j] if j < len(row) else None
            d[name] = _cell_value(v)
        out.append(d)
    return out


def _parse_excel(data: bytes) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheets = []
    primary = None
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            header = None
            used: set[str] = set()
            raw_rows = []
            for row in ws.iter_rows(values_only=True):
                if header is None:
                    cleaned = [str(c).strip() if c is not None else "" for c in row]
                    if sum(1 for c in cleaned if c) >= 2:
                        header = [_clean_header(c, used) for c in cleaned]
                    continue
                if not any(not _is_missing(v) for v in row):
                    continue
                raw_rows.append(list(row))
            sheets.append({"name": name, "rows": len(raw_rows), "cols": len(header or [])})
            if header is not None and (primary is None or not primary["rows"]):
                primary = {"columns": header, "rows": raw_rows}
    finally:
        wb.close()

    if primary is None:
        return {"columns": [], "rows": [], "sheets": sheets}
    rows = _rows_to_dicts(primary["columns"], primary["rows"])
    return {"columns": primary["columns"], "rows": rows, "sheets": sheets}


def _parse_delimited(data: bytes, delim: str) -> dict:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    header = None
    used: set[str] = set()
    raw_rows = []
    for row in reader:
        if header is None:
            if sum(1 for c in row if str(c).strip()) >= 2:
                header = [_clean_header(c, used) for c in row]
            continue
        if not any(str(c).strip() for c in row):
            continue
        raw_rows.append(row)
    if header is None:
        return {"columns": [], "rows": []}
    return {"columns": header, "rows": _rows_to_dicts(header, raw_rows)}


def _parse_json(data: bytes) -> dict:
    obj = json.loads(data.decode("utf-8-sig", errors="replace"))
    if isinstance(obj, dict):
        if "rows" in obj and isinstance(obj["rows"], list):
            obj = obj["rows"]
        elif "data" in obj and isinstance(obj["data"], list):
            obj = obj["data"]
        else:
            obj = [obj]
    if not isinstance(obj, list):
        return {"columns": [], "rows": []}
    rows = [r for r in obj if isinstance(r, dict)]
    used: set[str] = set()
    columns: list[str] = []
    for r in rows:
        for k in r.keys():
            if k not in used:
                used.add(k)
                columns.append(_clean_header(k, used))
    norm = [{c: _cell_value(r.get(c)) for c in columns} for r in rows]
    return {"columns": columns, "rows": norm}


def _parse_ndjson(data: bytes) -> dict:
    rows = []
    used: set[str] = set()
    columns: list[str] = []
    for line in data.decode("utf-8-sig", errors="replace").splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            r = json.loads(line)
        except ValueError:
            continue
        if not isinstance(r, dict):
            continue
        for k in r.keys():
            if k not in used:
                used.add(k)
                columns.append(_clean_header(k, used))
        rows.append(r)
    norm = [{c: _cell_value(r.get(c)) for c in columns} for r in rows]
    return {"columns": columns, "rows": norm}


def parse_table(data: bytes, filename: str) -> dict:
    ext = Path(filename).suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        return _parse_excel(data)
    if ext == ".csv":
        return _parse_delimited(data, ",")
    if ext == ".tsv":
        return _parse_delimited(data, "\t")
    if ext == ".json":
        return _parse_json(data)
    if ext in (".ndjson", ".jsonl"):
        return _parse_ndjson(data)
    # fallback sniff
    text = data.decode("utf-8-sig", errors="replace").lstrip()
    if text.startswith("[") or text.startswith("{"):
        return _parse_json(data)
    return _parse_delimited(data, ",")


# ---------------------------------------------------------------------------
# profiling
# ---------------------------------------------------------------------------
def _histogram(nums: list[float], bins: int = 8) -> list[dict]:
    if not nums:
        return []
    lo, hi = min(nums), max(nums)
    if hi == lo:
        return [{"bin": _fmt_num(lo), "count": len(nums)}]
    width = (hi - lo) / bins
    counts = [0] * bins
    for v in nums:
        idx = min(bins - 1, int((v - lo) / width))
        counts[idx] += 1
    out = []
    for i, c in enumerate(counts):
        a = lo + i * width
        b = lo + (i + 1) * width
        out.append({"bin": f"{_fmt_num(a)}–{_fmt_num(b)}", "count": c})
    return out


def _fmt_num(n: float) -> str:
    if abs(n) >= 1000:
        return f"{n:,.0f}"
    if n == int(n):
        return str(int(n))
    return f"{n:,.1f}"


def _numeric_stats(nums: list[float]) -> dict:
    nums = sorted(nums)
    n = len(nums)
    mean = sum(nums) / n
    sd = statistics.pstdev(nums) if n > 1 else 0.0
    return {
        "min": round(min(nums), 2),
        "max": round(max(nums), 2),
        "mean": round(mean, 2),
        "median": round(statistics.median(nums), 2),
        "sum": round(sum(nums), 2),
        "stddev": round(sd, 2),
        "histogram": _histogram(nums),
    }


def _top_values(non_null: list, k: int = 10) -> list[dict]:
    counts = Counter(_short(v) for v in non_null)
    total = len(non_null) or 1
    return [
        {"value": val, "count": cnt, "pct": round(cnt / total * 100, 1)}
        for val, cnt in counts.most_common(k)
    ]


def _name_tokens(name: str) -> set[str]:
    return {t for t in re.split(r"[^a-z0-9]+", (name or "").lower()) if t}


def _is_id_column(name: str) -> bool:
    return bool(_name_tokens(name) & ID_TOKENS)


def _profile_column(name: str, values: list) -> dict:
    n = len(values)
    non_null = [v for v in values if not _is_missing(v)]
    prof = {
        "name": name,
        "nulls": n - len(non_null),
        "null_pct": round((n - len(non_null)) / n * 100, 1) if n else 0,
        "distinct": len({str(v) for v in non_null}),
    }
    if not non_null:
        prof["type"] = "empty"
        return prof

    # Identifiers / keys (BOL #, tracking #, zip, PRO, reference …) are never
    # numeric measures — classify as categorical or text instead.
    if _is_id_column(name):
        if prof["distinct"] <= 30:
            prof["type"] = "categorical"
            prof["categorical"] = _top_values(non_null)
        else:
            prof["type"] = "text"
            prof["top"] = _top_values(non_null, k=5)
        return prof

    nums = [f for f in (_try_number(v) for v in non_null) if f is not None]
    if len(nums) >= max(2, 0.7 * len(non_null)):
        prof["type"] = "numeric"
        prof["numeric"] = _numeric_stats(nums)
        return prof

    dates = sum(1 for v in non_null if _try_date(v))
    if dates >= max(2, 0.6 * len(non_null)):
        prof["type"] = "date"
        parsed = [d for d in (_parse_date(v) for v in non_null) if d]
        if parsed:
            mn, mx = min(parsed), max(parsed)
            prof["date"] = {
                "min": mn.strftime("%Y-%m-%d"),
                "max": mx.strftime("%Y-%m-%d"),
                "span_days": (mx - mn).days,
            }
        return prof

    if all(str(v).strip().lower() in ("true", "false", "yes", "no", "y", "n", "1", "0") for v in non_null):
        prof["type"] = "boolean"
        return prof

    distinct = prof["distinct"]
    if distinct <= 30 or distinct <= 0.15 * max(len(non_null), 1) + 5:
        prof["type"] = "categorical"
        prof["categorical"] = _top_values(non_null)
    else:
        prof["type"] = "text"
        prof["top"] = _top_values(non_null, k=5)
    return prof


def _profile(columns: list[str], rows: list[dict]) -> dict:
    cols_prof = []
    total_cells = 0
    missing_cells = 0
    for name in columns:
        values = [r.get(name) for r in rows]
        p = _profile_column(name, values)
        cols_prof.append(p)
        total_cells += len(values)
        missing_cells += p["nulls"]
    completeness = round((1 - missing_cells / total_cells) * 100, 1) if total_cells else 0.0
    return {"columns": cols_prof, "completeness_pct": completeness}


# ---------------------------------------------------------------------------
# chart suggestions
# ---------------------------------------------------------------------------
def _numeric_score(p: dict) -> float:
    name = p["name"].lower()
    s = 0.0
    if any(h in name for h in MONEY_HINTS):
        s += 6
    elif any(h in name for h in QTY_HINTS):
        s += 4
    if p["distinct"] >= 5:
        s += 2
    else:
        s -= 1
    return s


def _cat_score(p: dict) -> float:
    name = p["name"].lower()
    tokens = _name_tokens(name)
    s = 0.0
    if any(h in name for h in DIMENSION_HINTS):
        s += 6
    if tokens & ID_TOKENS:
        s -= 6
    if tokens & LOW_VALUE_TOKENS:
        s -= 4
    d = p["distinct"]
    if 2 <= d <= 30:
        s += 3
    elif d <= 100:
        s += 1
    return s


def _date_score(p: dict) -> float:
    return sum(w for k, w in DATE_HINTS if k in p["name"].lower())


GROUP_HINTS = ("carrier", "city", "warehouse", "brand", "category", "customer", "receiver", "destination")


def _group_dimension(cats: list[dict]) -> dict | None:
    """Pick the best dimension for a money breakdown — prefer granular dims
    (carrier / city / warehouse) over coarse ones (status / type)."""
    if not cats:
        return None
    for p in cats:
        if any(h in p["name"].lower() for h in GROUP_HINTS) and p["distinct"] >= 3:
            return p
    candidates = [p for p in cats if p["distinct"] >= 3]
    if candidates:
        return max(candidates, key=lambda p: p["distinct"])
    return cats[0]


def _grouped(rows: list[dict], dim: str, measure: str, agg: str = "sum") -> list[dict]:
    buckets: dict[str, list[float]] = {}
    for r in rows:
        key = str(r.get(dim) if not _is_missing(r.get(dim)) else "(blank)")
        v = _try_number(r.get(measure))
        if v is None:
            continue
        buckets.setdefault(key, []).append(v)
    out = []
    for key, vals in buckets.items():
        if agg == "avg":
            value = sum(vals) / len(vals)
        elif agg == "min":
            value = min(vals)
        elif agg == "max":
            value = max(vals)
        else:
            value = sum(vals)
        out.append({"label": _short(key), "value": round(value, 1), "count": len(vals)})
    out.sort(key=lambda x: -x["value"])
    return out[:MAX_CHART_ITEMS]


def _time_series(rows: list[dict], date_col: str, measure: str | None) -> list[dict]:
    pairs = []
    for r in rows:
        d = _parse_date(r.get(date_col))
        if d is None:
            continue
        v = _try_number(r.get(measure)) if measure else None
        pairs.append((d, v))
    if not pairs:
        return []
    span = (max(d for d, _ in pairs) - min(d for d, _ in pairs)).days
    monthly = span > 62
    buckets: dict[str, dict] = {}
    for d, v in pairs:
        key = d.strftime("%Y-%m") if monthly else d.strftime("%Y-%m-%d")
        b = buckets.setdefault(key, {"label": key, "count": 0, "sum": 0.0})
        b["count"] += 1
        if v is not None:
            b["sum"] += v
    out = []
    for key in sorted(buckets):
        b = buckets[key]
        value = b["sum"] if measure else b["count"]
        out.append({"label": key, "value": round(value, 1), "count": b["count"]})
    return out[-MAX_CHART_ITEMS:]


def _suggest_charts(columns: list[str], profs: list[dict], rows: list[dict]) -> list[dict]:
    numeric = sorted([p for p in profs if p["type"] == "numeric"], key=_numeric_score, reverse=True)
    cat = sorted(
        [p for p in profs if p["type"] == "categorical" and p["distinct"] >= 2],
        key=_cat_score, reverse=True,
    )
    dates = sorted([p for p in profs if p["type"] == "date"], key=_date_score, reverse=True)
    charts: list[dict] = []

    for p in cat[:3]:
        if _cat_score(p) <= 0:
            continue
        charts.append({
            "type": "bar",
            "title": f"Count by {p['name']}",
            "dimension": p["name"],
            "data": [{"label": t["value"], "value": t["count"]} for t in p["categorical"][:MAX_CHART_ITEMS]],
        })

    for p in numeric[:2]:
        if _numeric_score(p) <= 0:
            continue
        charts.append({
            "type": "histogram",
            "title": f"Distribution of {p['name']}",
            "column": p["name"],
            "data": p["numeric"]["histogram"],
        })

    if dates and rows:
        dc = dates[0]
        measure = numeric[0]["name"] if numeric else None
        data = _time_series(rows, dc["name"], measure)
        if data:
            charts.append({
                "type": "timeseries",
                "title": f"Trend over {dc['name']}",
                "column": dc["name"],
                "measure": measure,
                "data": data,
            })

    if cat and numeric:
        gd = _group_dimension(cat)
        measure = numeric[0]["name"]
        if gd:
            data = _grouped(rows, gd["name"], measure, "sum")
            if data:
                charts.append({
                    "type": "grouped",
                    "title": f"Total {measure} by {gd['name']}",
                    "dimension": gd["name"],
                    "measure": measure,
                    "data": data,
                })

    return charts


# ---------------------------------------------------------------------------
# dataset build / store
# ---------------------------------------------------------------------------
def _build_dataset(filename: str, table: dict) -> dict:
    columns = table["columns"][:MAX_COLS_STORE]
    rows = table["rows"][:MAX_ROWS_STORE]
    truncated = 1 if len(table["rows"]) > MAX_ROWS_STORE else 0
    profile = _profile(columns, rows)
    suggested = _suggest_charts(columns, profile["columns"], rows)
    name = Path(filename).stem or "dataset"
    return {
        "name": name,
        "filename": filename,
        "columns": columns,
        "row_count": len(rows),
        "col_count": len(columns),
        "truncated": truncated,
        "profile": profile,
        "suggested": suggested,
        "sheets": table.get("sheets") or [],
        "rows": rows,
    }


def _store(ds: dict) -> int:
    conn = storage._conn()
    cur = conn.execute(
        "INSERT INTO insights_datasets (name, filename, columns, profile, rows_json, "
        "row_count, col_count, truncated, created_at) VALUES (?,?,?,?,?,?,?,?,?)",
        (
            ds["name"],
            ds["filename"],
            json.dumps(ds["columns"]),
            json.dumps({"columns": ds["profile"]["columns"], "completeness_pct": ds["profile"]["completeness_pct"]}),
            json.dumps(ds["rows"], ensure_ascii=False, default=str),
            ds["row_count"],
            ds["col_count"],
            ds["truncated"],
            storage.now_iso(),
        ),
    )
    conn.commit()
    return cur.lastrowid


def _load_row(row) -> dict:
    d = dict(row)
    d["columns"] = json.loads(d.get("columns") or "[]")
    return d


# ---------------------------------------------------------------------------
# AI summary
# ---------------------------------------------------------------------------
SUMMARIZE_SYSTEM = """You are a data analyst inside Conductor, a desktop app for an e-commerce and
logistics operator. Given a profile and sample rows of a tabular dataset, write a
concise executive summary of what the data shows.

Respond with STRICT JSON ONLY — a single object:
{
  "summary": "<2-3 sentence plain-English overview of the dataset and its headline numbers>",
  "insights": ["<one key insight per string, 4-7 of them, each specific and grounded in the numbers>"],
  "recommendations": ["<one concrete next action per string, 2-4 of them>"]
}

RULES:
- Be concrete: cite real figures (totals, top categories, ranges, trends) from the data.
- Flag outliers, missing data, and unusual patterns when present.
- Do not invent figures that are not in the provided profile/sample.
- Only output the JSON object — no prose, no markdown."""  # noqa: E501


def _build_prompt(ds: dict) -> str:
    lines = [f"Dataset: {ds['name']} ({ds['row_count']} rows, {ds['col_count']} columns)"]
    lines.append("Column profile:")
    for p in ds["profile"]["columns"][:25]:
        t = p["type"]
        desc = f"  - {p['name']} [{t}; {p['distinct']} distinct; {p['null_pct']}% missing]"
        if t == "numeric":
            n = p["numeric"]
            desc += f" min={n['min']} max={n['max']} mean={n['mean']} sum={n['sum']}"
        elif t == "categorical":
            tops = ", ".join(f"{x['value']}={x['pct']}%" for x in p["categorical"][:6])
            desc += f" top: {tops}"
        elif t == "date":
            d = p.get("date") or {}
            desc += f" range: {d.get('min')} .. {d.get('max')}"
        lines.append(desc)
    lines.append("Sample rows:")
    for r in ds["rows"][:8]:
        compact = {k: _short(v, 80) for k, v in r.items() if not _is_missing(v)}
        lines.append("  " + json.dumps(compact, ensure_ascii=False, default=str))
    return "\n".join(lines)


def _provider_ready():
    try:
        from ai_ingest import _provider_ready as ready
        return ready()
    except Exception:
        return None


# ---------------------------------------------------------------------------
# routes
# ---------------------------------------------------------------------------
@router.post("/upload")
async def upload(file: UploadFile = File(...)):
    ext = Path(file.filename or "").suffix.lower()
    if ext not in SUPPORTED_EXTS:
        raise HTTPException(400, f"Unsupported file type '{ext or '?'}' — use xlsx/xlsm/csv/tsv/json/ndjson")
    data = await file.read()
    if len(data) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, "File too large (max 50 MB)")
    try:
        table = parse_table(data, file.filename or "data.csv")
    except Exception as exc:
        raise HTTPException(400, f"Could not parse file: {exc}")
    if not table["rows"] or not table["columns"]:
        raise HTTPException(400, "No tabular data found in this file")
    ds = _build_dataset(file.filename or "data.csv", table)
    ds_id = _store(ds)
    return {
        "id": ds_id,
        "name": ds["name"],
        "filename": ds["filename"],
        "row_count": ds["row_count"],
        "col_count": ds["col_count"],
        "truncated": ds["truncated"],
        "sheets": ds["sheets"],
        "profile": ds["profile"],
        "suggested": ds["suggested"],
    }


@router.get("")
def list_datasets():
    rows = storage._conn().execute(
        "SELECT id, name, filename, row_count, col_count, truncated, created_at "
        "FROM insights_datasets ORDER BY id DESC LIMIT 100"
    ).fetchall()
    return {"datasets": [dict(r) for r in rows]}


@router.get("/{dataset_id}")
def get_dataset(dataset_id: int):
    row = storage._conn().execute(
        "SELECT * FROM insights_datasets WHERE id=?", (dataset_id,)
    ).fetchone()
    if not row:
        raise HTTPException(404, "Dataset not found")
    d = _load_row(row)
    try:
        rows = json.loads(d["rows_json"])
    except Exception:
        rows = []
    try:
        profile = json.loads(d["profile"])
    except Exception:
        profile = {"columns": [], "completeness_pct": 0}
    cols_prof = profile.get("columns", [])
    suggested = _suggest_charts(d["columns"], cols_prof, rows)
    return {
        "id": d["id"],
        "name": d["name"],
        "filename": d["filename"],
        "columns": d["columns"],
        "row_count": d["row_count"],
        "col_count": d["col_count"],
        "truncated": d["truncated"],
        "created_at": d["created_at"],
        "completeness_pct": profile.get("completeness_pct", 0),
        "profile": cols_prof,
        "suggested": suggested,
        "rows": rows[:MAX_PREVIEW],
        "preview_count": min(len(rows), MAX_PREVIEW),
    }


@router.get("/{dataset_id}/rows")
def get_rows(dataset_id: int, offset: int = 0, limit: int = 200, q: str = ""):
    row = storage._conn().execute(
        "SELECT rows_json, row_count FROM insights_datasets WHERE id=?", (dataset_id,)
    ).fetchone()
    if not row:
        raise HTTPException(404, "Dataset not found")
    try:
        rows = json.loads(row["rows_json"])
    except Exception:
        rows = []
    if q:
        ql = q.lower()
        rows = [r for r in rows if any(ql in str(v).lower() for v in r.values())]
    total = len(rows)
    page = rows[offset:offset + min(max(limit, 1), 1000)]
    return {"total": total, "offset": offset, "rows": page}


@router.post("/{dataset_id}/summarize")
def summarize(dataset_id: int):
    row = storage._conn().execute(
        "SELECT * FROM insights_datasets WHERE id=?", (dataset_id,)
    ).fetchone()
    if not row:
        raise HTTPException(404, "Dataset not found")
    d = _load_row(row)
    try:
        rows = json.loads(d["rows_json"])
    except Exception:
        rows = []
    try:
        profile = json.loads(d["profile"])
    except Exception:
        profile = {"columns": [], "completeness_pct": 0}
    ds = {
        "name": d["name"],
        "row_count": d["row_count"],
        "col_count": d["col_count"],
        "profile": {"columns": profile.get("columns", [])},
        "rows": rows,
    }

    ready = _provider_ready()
    if not ready:
        return {
            "ai": False,
            "error": "No AI provider key configured — add one in Settings → AI Chat, then summarise again.",
            "summary": None,
        }

    provider, model, api_key = ready
    import providers

    messages = [
        {"role": "system", "content": SUMMARIZE_SYSTEM},
        {"role": "user", "content": _build_prompt(ds)},
    ]
    try:
        text = ""
        for ev in providers.stream_provider(provider, messages, model=model, api_key=api_key):
            if ev["type"] == "text":
                text += ev["text"]
            elif ev["type"] == "error":
                raise ValueError(f"Provider error: {ev.get('code')} {ev.get('message')}")
        start, end = text.find("{"), text.rfind("}")
        if start == -1 or end <= start:
            raise ValueError("AI returned unparseable output (expected JSON)")
        parsed = json.loads(text[start:end + 1])
        if not isinstance(parsed, dict):
            raise ValueError("AI returned an unexpected shape")
        return {
            "ai": True,
            "error": None,
            "summary": {
                "summary": str(parsed.get("summary") or "").strip(),
                "insights": [str(x) for x in (parsed.get("insights") or [])],
                "recommendations": [str(x) for x in (parsed.get("recommendations") or [])],
            },
        }
    except Exception as exc:
        return {"ai": False, "error": str(exc), "summary": None}


@router.delete("/{dataset_id}", status_code=204)
def delete_dataset(dataset_id: int):
    conn = storage._conn()
    cur = conn.execute("DELETE FROM insights_datasets WHERE id=?", (dataset_id,))
    conn.commit()
    if cur.rowcount == 0:
        raise HTTPException(404, "Dataset not found")
    return None


@router.get("/{dataset_id}/export")
def export_csv(dataset_id: int):
    row = storage._conn().execute(
        "SELECT name, columns, rows_json FROM insights_datasets WHERE id=?", (dataset_id,)
    ).fetchone()
    if not row:
        raise HTTPException(404, "Dataset not found")
    columns = json.loads(row["columns"] or "[]")
    try:
        rows = json.loads(row["rows_json"])
    except Exception:
        rows = []
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(columns)
    for r in rows:
        writer.writerow(["" if _is_missing(r.get(c)) else r.get(c) for c in columns])
    name = re.sub(r"[^A-Za-z0-9._-]+", "_", row["name"] or "dataset")
    return Response(
        content=buf.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{name}.csv"'},
    )
