"""KPI Engine, Employee Performance Evaluation & DataWrangler backend router.

Provides storage and endpoints for:
- KPI metrics and monthly entry tracking (seeded from Global KPIs Excel).
- Employee performance evaluation scorecards (% to goal, composite scores).
- Natural Language Processing (NLP) to KPI conversion.
- DataWrangler dataset persistence & transformation rules.
"""
from __future__ import annotations

import json
import math
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import APIRouter, HTTPException, Query
import storage

ATTACHMENT_EXCEL = Path(
    r"C:\Users\GabeMaher\AppData\Local\hermes\profiles\sammy\attachments\Global KPIs (1).xlsx"
)

kpi_router = APIRouter(prefix="/api/kpis", tags=["kpi-performance"])
wrangler_router = APIRouter(prefix="/api/wrangler", tags=["data-wrangler"])


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def init_kpi_db() -> None:
    conn = storage._conn()
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS kpi_metrics (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            department TEXT NOT NULL,
            owner TEXT NOT NULL,
            kpi_name TEXT NOT NULL,
            expected_value REAL,
            metric_type TEXT DEFAULT '%',
            weight REAL DEFAULT 1.0,
            formula TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS kpi_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            kpi_id INTEGER NOT NULL,
            period_date TEXT NOT NULL,
            actual_value REAL,
            sparkline_val TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(kpi_id) REFERENCES kpi_metrics(id) ON DELETE CASCADE,
            UNIQUE(kpi_id, period_date)
        );
        CREATE TABLE IF NOT EXISTS wrangled_datasets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            source TEXT DEFAULT 'manual',
            columns TEXT NOT NULL,
            rows TEXT NOT NULL,
            transformations TEXT DEFAULT '[]',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_kpi_metrics_dept ON kpi_metrics(department);
        CREATE INDEX IF NOT EXISTS idx_kpi_metrics_owner ON kpi_metrics(owner);
        CREATE INDEX IF NOT EXISTS idx_kpi_entries_kpi ON kpi_entries(kpi_id);
        """
    )
    conn.commit()


def seed_kpis_from_excel(force: bool = False) -> int:
    """Seed kpi_metrics and kpi_entries from Global KPIs (1).xlsx 'MAIN' sheet."""
    init_kpi_db()
    conn = storage._conn()
    existing_count = conn.execute("SELECT COUNT(*) FROM kpi_metrics").fetchone()[0]
    if existing_count > 0 and not force:
        return existing_count

    if not ATTACHMENT_EXCEL.exists():
        return 0

    try:
        import openpyxl
    except ImportError:
        return 0

    try:
        wb = openpyxl.load_workbook(str(ATTACHMENT_EXCEL), data_only=True)
        sheet = wb["MAIN"] if "MAIN" in wb.sheetnames else wb.active
    except Exception:
        return 0

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return 0

    # Header row contains dates starting at index 6
    header = [str(c) if c is not None else "" for c in rows[0]]
    date_cols: list[tuple[int, str]] = []
    for col_idx in range(6, len(header)):
        val = header[col_idx].strip()
        if val:
            # Parse ISO date string prefix e.g. 2026-05-01
            d_match = re.search(r"\d{4}-\d{2}-\d{2}", val)
            if d_match:
                date_cols.append((col_idx, d_match.group(0)))

    current_dept = "General"
    inserted_metrics = 0

    for row_idx in range(1, len(rows)):
        r = rows[row_idx]
        if not any(r):
            continue

        dept_val = str(r[0]).strip() if r[0] is not None else ""
        owner_val = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""
        kpi_val = str(r[2]).strip() if len(r) > 2 and r[2] is not None else ""
        exp_val_raw = r[3] if len(r) > 3 else None
        mtype_val = str(r[4]).strip() if len(r) > 4 and r[4] is not None else "%"

        # Check for Department section header rows
        if owner_val and not kpi_val and not exp_val_raw and owner_val == dept_val:
            current_dept = owner_val
            continue
        elif dept_val and not owner_val and not kpi_val:
            current_dept = dept_val
            continue

        if not kpi_val or not owner_val:
            continue

        department = current_dept if current_dept else "General"

        # Parse expected value
        exp_val: float | None = None
        if exp_val_raw is not None:
            try:
                exp_val = float(exp_val_raw)
            except (ValueError, TypeError):
                exp_val = None

        cur = conn.execute(
            """INSERT INTO kpi_metrics (department, owner, kpi_name, expected_value, metric_type, weight, created_at)
               VALUES (?, ?, ?, ?, ?, 1.0, ?)""",
            (department, owner_val, kpi_val, exp_val, mtype_val or "%", now_iso()),
        )
        kpi_id = cur.lastrowid
        inserted_metrics += 1

        # Process monthly entries
        for col_idx, period_date in date_cols:
            if col_idx < len(r):
                act_raw = r[col_idx]
                if act_raw is not None:
                    act_str = str(act_raw).strip().rstrip("*")
                    try:
                        act_val = float(act_str)
                        conn.execute(
                            """INSERT OR REPLACE INTO kpi_entries
                               (kpi_id, period_date, actual_value, sparkline_val, created_at, updated_at)
                               VALUES (?, ?, ?, ?, ?, ?)""",
                            (kpi_id, period_date, act_val, str(act_raw), now_iso(), now_iso()),
                        )
                    except (ValueError, TypeError):
                        pass

    conn.commit()
    return inserted_metrics


@kpi_router.get("")
def list_kpis(
    department: str | None = None,
    owner: str | None = None,
    search: str | None = None,
) -> list[dict[str, Any]]:
    init_kpi_db()
    conn = storage._conn()
    query = "SELECT * FROM kpi_metrics WHERE 1=1"
    params: list[Any] = []

    if department:
        query += " AND department = ?"
        params.append(department)
    if owner:
        query += " AND owner = ?"
        params.append(owner)
    if search:
        query += " AND (kpi_name LIKE ? OR department LIKE ? OR owner LIKE ?)"
        term = f"%{search}%"
        params.extend([term, term, term])

    query += " ORDER BY department, owner, kpi_name"
    rows = conn.execute(query, params).fetchall()

    metrics = []
    for r in rows:
        m = dict(r)
        # Fetch entries
        entries = conn.execute(
            "SELECT period_date, actual_value, sparkline_val FROM kpi_entries WHERE kpi_id = ? ORDER BY period_date DESC",
            (m["id"],),
        ).fetchall()
        m["entries"] = [dict(e) for e in entries]
        m["latest_entry"] = m["entries"][0] if m["entries"] else None
        metrics.append(m)

    return metrics


@kpi_router.post("/seed")
def seed_kpis(force: bool = False) -> dict[str, Any]:
    count = seed_kpis_from_excel(force=force)
    return {"ok": True, "count": count, "message": f"Seeded {count} KPI metrics from Excel"}


@kpi_router.post("")
def create_kpi(body: dict[str, Any]) -> dict[str, Any]:
    init_kpi_db()
    department = str(body.get("department") or "General").strip()
    owner = str(body.get("owner") or "Unassigned").strip()
    kpi_name = str(body.get("kpi_name") or "").strip()
    if not kpi_name:
        raise HTTPException(400, "kpi_name is required")

    exp_val = body.get("expected_value")
    exp_val_float = float(exp_val) if exp_val is not None and str(exp_val).strip() != "" else None
    metric_type = str(body.get("metric_type") or "%").strip()
    weight = float(body.get("weight") or 1.0)
    formula = str(body.get("formula") or "").strip()
    notes = str(body.get("notes") or "").strip()

    conn = storage._conn()
    cur = conn.execute(
        """INSERT INTO kpi_metrics (department, owner, kpi_name, expected_value, metric_type, weight, formula, notes, created_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (department, owner, kpi_name, exp_val_float, metric_type, weight, formula, notes, now_iso()),
    )
    conn.commit()
    kpi_id = cur.lastrowid
    return {"ok": True, "id": kpi_id, "message": "KPI metric created successfully"}


@kpi_router.delete("/{kpi_id}")
def delete_kpi(kpi_id: int) -> dict[str, Any]:
    init_kpi_db()
    conn = storage._conn()
    conn.execute("DELETE FROM kpi_metrics WHERE id = ?", (kpi_id,))
    conn.execute("DELETE FROM kpi_entries WHERE kpi_id = ?", (kpi_id,))
    conn.commit()
    return {"ok": True, "message": f"Deleted KPI metric {kpi_id}"}


@kpi_router.post("/entries")
def upsert_kpi_entry(body: dict[str, Any]) -> dict[str, Any]:
    init_kpi_db()
    kpi_id = int(body.get("kpi_id") or 0)
    period_date = str(body.get("period_date") or "").strip()
    if not kpi_id or not period_date:
        raise HTTPException(400, "kpi_id and period_date are required")

    actual_val = body.get("actual_value")
    act_float = float(actual_val) if actual_val is not None and str(actual_val).strip() != "" else None
    sparkline_val = str(body.get("sparkline_val") or "").strip()

    conn = storage._conn()
    conn.execute(
        """INSERT OR REPLACE INTO kpi_entries (kpi_id, period_date, actual_value, sparkline_val, created_at, updated_at)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (kpi_id, period_date, act_float, sparkline_val, now_iso(), now_iso()),
    )
    conn.commit()
    return {"ok": True, "message": "KPI entry recorded"}


@kpi_router.get("/employee-evaluation")
def evaluate_employees() -> dict[str, Any]:
    """Calculate employee performance scorecards per owner."""
    init_kpi_db()
    # Ensure seeded if empty
    seed_kpis_from_excel(force=False)

    conn = storage._conn()
    metrics = conn.execute("SELECT * FROM kpi_metrics").fetchall()

    evaluations: dict[str, list[dict[str, Any]]] = {}

    for m_row in metrics:
        m = dict(m_row)
        owner = m["owner"]
        entries = conn.execute(
            "SELECT * FROM kpi_entries WHERE kpi_id = ? ORDER BY period_date DESC", (m["id"],)
        ).fetchall()
        latest_entry = entries[0] if entries else None

        actual = latest_entry["actual_value"] if latest_entry else None
        expected = m["expected_value"]
        mtype = m["metric_type"]

        pct_to_goal: float | None = None
        status = "neutral"

        if actual is not None and expected is not None and expected != 0:
            # Lower is better for days/overdue
            if "time" in m["kpi_name"].lower() or "days" in mtype.lower() or "overdue" in m["kpi_name"].lower():
                pct_to_goal = round((expected / actual) * 100, 1) if actual != 0 else 100.0
            else:
                pct_to_goal = round((actual / expected) * 100, 1)

            if pct_to_goal >= 100:
                status = "exceeds"
            elif pct_to_goal >= 85:
                status = "meets"
            else:
                status = "needs_improvement"

        eval_item = {
            "kpi_id": m["id"],
            "kpi_name": m["kpi_name"],
            "department": m["department"],
            "owner": owner,
            "metric_type": mtype,
            "expected_value": expected,
            "actual_value": actual,
            "pct_to_goal": pct_to_goal,
            "status": status,
            "weight": m["weight"],
            "period_date": latest_entry["period_date"] if latest_entry else None,
            "history": [
                {"period": e["period_date"], "value": e["actual_value"]}
                for e in entries[:6]
            ],
        }

        evaluations.setdefault(owner, []).append(eval_item)

    scorecards = []
    for owner, owner_metrics in evaluations.items():
        valid_pcts = [item["pct_to_goal"] for item in owner_metrics if item["pct_to_goal"] is not None]
        avg_score = round(sum(valid_pcts) / len(valid_pcts), 1) if valid_pcts else None

        rating = "Satisfactory"
        if avg_score is not None:
            if avg_score >= 105:
                rating = "Exceptional"
            elif avg_score >= 95:
                rating = "Exceeds Expectations"
            elif avg_score >= 85:
                rating = "Meets Expectations"
            else:
                rating = "Needs Focus"

        scorecards.append(
            {
                "owner": owner,
                "total_kpis": len(owner_metrics),
                "composite_score": avg_score,
                "performance_rating": rating,
                "metrics": owner_metrics,
            }
        )

    scorecards.sort(key=lambda x: (x["composite_score"] or 0), reverse=True)
    return {"ok": True, "scorecards": scorecards, "total_owners": len(scorecards)}


@kpi_router.post("/nlp-convert")
def nlp_to_kpi(body: dict[str, Any]) -> dict[str, Any]:
    """Parse natural language prompts into KPI metrics or performance reports."""
    prompt = str(body.get("prompt") or "").strip()
    if not prompt:
        raise HTTPException(400, "prompt is required")

    prompt_lower = prompt.lower()
    init_kpi_db()

    # Intent 1: Employee Performance Evaluation Request
    if any(k in prompt_lower for k in ("eval", "employee", "performance", "scorecard", "report", "review")):
        scorecard_data = evaluate_employees()
        scorecards = scorecard_data.get("scorecards", [])

        # Filter by owner if mentioned
        owner_match = None
        for sc in scorecards:
            if sc["owner"].lower() in prompt_lower:
                owner_match = sc
                break

        if owner_match:
            summary = (
                f"**Employee Performance Report: {owner_match['owner']}**\n"
                f"- **Composite Performance Score:** {owner_match['composite_score']}%\n"
                f"- **Overall Rating:** {owner_match['performance_rating']}\n"
                f"- **Tracked KPIs:** {owner_match['total_kpis']}\n\n"
                f"**Key Metrics Summary:**\n"
            )
            for m in owner_match["metrics"]:
                act = m["actual_value"] if m["actual_value"] is not None else "N/A"
                exp = m["expected_value"] if m["expected_value"] is not None else "N/A"
                pct = f"{m['pct_to_goal']}%" if m["pct_to_goal"] is not None else "N/A"
                summary += f"- **{m['kpi_name']}** ({m['department']}): Actual `{act}` vs Expected `{exp}` ({pct} to goal)\n"
            return {
                "ok": True,
                "type": "performance_report",
                "owner": owner_match["owner"],
                "summary": summary,
                "data": owner_match,
            }
        else:
            summary = "**Company-Wide Employee Performance Overview**\n\n"
            for sc in scorecards:
                summary += f"- **{sc['owner']}**: Score `{sc['composite_score']}%` ({sc['performance_rating']}) across {sc['total_kpis']} KPIs\n"
            return {
                "ok": True,
                "type": "team_performance_report",
                "summary": summary,
                "data": scorecards,
            }

    # Intent 2: Create new KPI metric from prompt
    dept = "Operations"
    if "catalog" in prompt_lower:
        dept = "Catalog"
    elif "case" in prompt_lower:
        dept = "Cases"
    elif "fba" in prompt_lower:
        dept = "FBA"

    owner = "Gabe"
    for name in ("gabe", "alice", "carlos", "jelena", "francis"):
        if name in prompt_lower:
            owner = name.title()
            break

    # Extract target number if mentioned e.g. "target 95%", "expected 10"
    num_match = re.search(r"(?:target|expected|goal|of)\s*(\d+(?:\.\d+)?)", prompt_lower)
    exp_val = float(num_match.group(1)) if num_match else 100.0

    mtype = "%"
    if "days" in prompt_lower or "day" in prompt_lower:
        mtype = "Days"
    elif "count" in prompt_lower or "#" in prompt:
        mtype = "#"
    elif "$" in prompt or "dollar" in prompt_lower:
        mtype = "$"

    kpi_name = prompt.capitalize()
    if len(kpi_name) > 60:
        kpi_name = kpi_name[:57] + "..."

    # Create the KPI
    created = create_kpi(
        {
            "department": dept,
            "owner": owner,
            "kpi_name": kpi_name,
            "expected_value": exp_val,
            "metric_type": mtype,
            "weight": 1.0,
            "formula": "NLP Auto-generated metric",
            "notes": f"Generated from NLP prompt: {prompt}",
        }
    )

    return {
        "ok": True,
        "type": "kpi_created",
        "summary": f"Created new KPI **{kpi_name}** assigned to **{owner}** ({dept}) with target **{exp_val} {mtype}**.",
        "kpi": created,
    }


# --- DataWrangler Endpoints ---

@wrangler_router.get("/datasets")
def list_wrangled_datasets() -> list[dict[str, Any]]:
    init_kpi_db()
    conn = storage._conn()
    rows = conn.execute("SELECT id, name, source, columns, created_at, updated_at FROM wrangled_datasets ORDER BY id DESC").fetchall()
    result = []
    for r in rows:
        item = dict(r)
        try:
            item["columns"] = json.loads(item["columns"])
        except Exception:
            item["columns"] = []
        result.append(item)
    return result


@wrangler_router.get("/datasets/{dataset_id}")
def get_wrangled_dataset(dataset_id: int) -> dict[str, Any]:
    init_kpi_db()
    conn = storage._conn()
    row = conn.execute("SELECT * FROM wrangled_datasets WHERE id = ?", (dataset_id,)).fetchone()
    if not row:
        raise HTTPException(404, "Dataset not found")
    item = dict(row)
    for key in ("columns", "rows", "transformations"):
        try:
            item[key] = json.loads(item[key])
        except Exception:
            item[key] = []
    return item


@wrangler_router.post("/datasets")
def save_wrangled_dataset(body: dict[str, Any]) -> dict[str, Any]:
    init_kpi_db()
    name = str(body.get("name") or "Wrangled Dataset").strip()
    source = str(body.get("source") or "manual").strip()
    columns = body.get("columns") or []
    rows = body.get("rows") or []
    transformations = body.get("transformations") or []

    conn = storage._conn()
    cur = conn.execute(
        """INSERT INTO wrangled_datasets (name, source, columns, rows, transformations, created_at, updated_at)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (
            name,
            source,
            json.dumps(columns),
            json.dumps(rows),
            json.dumps(transformations),
            now_iso(),
            now_iso(),
        ),
    )
    conn.commit()
    return {"ok": True, "id": cur.lastrowid, "name": name, "message": "Dataset saved successfully"}


@wrangler_router.post("/transform")
def transform_dataset(body: dict[str, Any]) -> dict[str, Any]:
    """Execute server-side wrangling transformations on a dataset payload."""
    columns: list[str] = body.get("columns") or []
    rows: list[dict[str, Any]] = body.get("rows") or []
    ops: list[dict[str, Any]] = body.get("transformations") or []

    out_rows = [dict(r) for r in rows]
    out_cols = list(columns)

    for op in ops:
        kind = op.get("kind")
        col = op.get("column")

        if kind == "filter" and col in out_cols:
            val = str(op.get("value") or "").lower()
            mode = op.get("mode", "contains")
            if mode == "contains":
                out_rows = [r for r in out_rows if val in str(r.get(col, "")).lower()]
            elif mode == "equals":
                out_rows = [r for r in out_rows if str(r.get(col, "")).lower() == val]
            elif mode == "not_empty":
                out_rows = [r for r in out_rows if str(r.get(col, "")).strip() != ""]

        elif kind == "fill_null" and col in out_cols:
            fill = op.get("fill_value", "N/A")
            for r in out_rows:
                if r.get(col) is None or str(r.get(col)).strip() == "":
                    r[col] = fill

        elif kind == "add_formula":
            new_col = str(op.get("new_column") or "Computed").strip()
            formula = str(op.get("formula") or "").strip()  # e.g. "colA / colB" or "colA * 100"
            if new_col and new_col not in out_cols:
                out_cols.append(new_col)

            for r in out_rows:
                # Basic formula evaluator replacing column names with numeric values
                expr = formula
                for c in out_cols:
                    if c in expr:
                        raw = r.get(c, 0)
                        try:
                            num = float(raw)
                        except (ValueError, TypeError):
                            num = 0.0
                        expr = expr.replace(c, str(num))
                try:
                    # Safe limited eval for basic operations
                    val = eval(expr, {"__builtins__": None, "math": math}, {})
                    r[new_col] = round(val, 4) if isinstance(val, (int, float)) else str(val)
                except Exception:
                    r[new_col] = 0.0

    return {
        "ok": True,
        "columns": out_cols,
        "rows": out_rows,
        "row_count": len(out_rows),
        "col_count": len(out_cols),
    }
