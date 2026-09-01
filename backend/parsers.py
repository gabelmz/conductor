"""Format-aware catalog parsers for real-world e-commerce exports and files.

Supports:
  * Excel (.xlsx, .xlsm, .xlsb, .xls)
  * Delimited text (.csv, .tsv, .tab, .txt) with auto-delimiter and multi-encoding support
  * Markdown tables (.md, .markdown)
  * Word tables (.docx)
  * PDF tables (.pdf)
  * JSON / NDJSON / JSONL
"""
from __future__ import annotations

import csv
import io
import json
import re
from pathlib import Path

# --------------------------------------------------------------------------
# Amazon root category -> compliance category
# --------------------------------------------------------------------------
AMAZON_CATEGORY_MAP = [
    (re.compile(r"electronics|computers|cell phones|video games|home audio|television", re.I), "electronics"),
    (re.compile(r"toys|games|baby|kids|children", re.I), "toys"),
    (re.compile(r"clothing|shoes|jewelry|apparel|fashion|accessories|watches|handbags|luggage", re.I), "textile"),
    (re.compile(r"beauty|personal care|cosmetic|skin care|hair care|makeup|fragrance", re.I), "cosmetics"),
    (re.compile(r"health|household|supplement|vitamin|nutrition|wellness|personal health", re.I), "general"),
    (re.compile(r"kitchen|dining|cookware|food service|restaurant", re.I), "kitchen"),
    (re.compile(r"grocery|food|beverage|pantry", re.I), "food-contact"),
    (re.compile(r"home|furniture|garden|tools|sports|outdoor", re.I), "general"),
]

ITEMTYPE_MAP = [
    (re.compile(r"battery|power bank|charger|adapter|cable|electronic|electric", re.I), "electronics"),
    (re.compile(r"wireless|bluetooth|wifi|radio|rf|transmitter|headset|earbud|smart", re.I), "wireless"),
    (re.compile(r"toy|game|puzzle|plush|doll|playset", re.I), "toys"),
    (re.compile(r"cosmetic|makeup|serum|lotion|cream|shampoo|conditioner|soap|fragrance|deodorant", re.I), "cosmetics"),
    (re.compile(r"shirt|pants|dress|jacket|sock|underwear|sweater|textile|fabric|apparel", re.I), "textile"),
    (re.compile(r"supplement|vitamin|herbal|probiotic|protein|nutritional", re.I), "general"),
    (re.compile(r"cup|mug|bowl|plate|bottle|container|cookware|pan|utensil", re.I), "food-contact"),
    (re.compile(r"brush|comb|hair accessory|hair clip", re.I), "personal-care"),
]

WIRELESS_HINTS = re.compile(r"wireless|bluetooth|wifi|wi-fi|rf |radio|2\.4ghz|5ghz|zigbee|nfc", re.I)
BATTERY_HINTS = re.compile(r"batter|rechargeab|li-ion|lithium|usb-c|usb c|power bank", re.I)
HAZMAT_HINTS = re.compile(r"hazmat|hazardous|flammable|corrosive|toxic|dangerous goods|class 9", re.I)


def infer_category(category_str: str, item_type: str = "", title: str = "") -> str:
    text = f"{category_str} {item_type} {title}"
    for pattern, cat in AMAZON_CATEGORY_MAP:
        if pattern.search(text):
            return cat
    for pattern, cat in ITEMTYPE_MAP:
        if pattern.search(text):
            return cat
    return "general"


def infer_attributes(row: dict) -> dict:
    attrs: dict = {}
    title = str(row.get("title", "") or "")

    batt_req = str(row.get("batteries_required", row.get("battery", "")) or "").strip().lower()
    batt_inc = str(row.get("batteries_included", "") or "").strip().lower()
    has_battery_col = batt_req in ("yes", "true", "1", "y") or batt_inc in ("yes", "true", "1", "y")
    if has_battery_col or BATTERY_HINTS.search(title):
        attrs["battery"] = True
        attrs["battery_type"] = "li-ion" if re.search(r"li-ion|lithium|li-po", title, re.I) else "unknown"
        if batt_req in ("yes", "true", "1", "y"):
            attrs["batteries_required"] = True
    else:
        attrs["battery"] = False

    if WIRELESS_HINTS.search(title) or str(row.get("wireless", "") or "").strip().lower() in ("yes", "true", "1"):
        attrs["wireless"] = True
    else:
        attrs["wireless"] = False

    hazmat = str(row.get("is_hazmat", row.get("hazardous_materials", "")) or "").strip().lower()
    substances: list[str] = []
    if hazmat in ("yes", "true", "1", "y") or HAZMAT_HINTS.search(hazmat):
        substances.append("hazmat")
    for token in ("lead", "cadmium", "mercury", "phthalate", "bpa", "nickel", "asbestos", "formaldehyde"):
        if re.search(token, title, re.I) or re.search(token, hazmat, re.I):
            substances.append(token)
    attrs["restricted_substances"] = substances

    materials = row.get("material", "")
    if materials:
        parts = [m.strip() for m in str(materials).replace(";", ",").split(",") if m.strip()]
        if parts:
            attrs["materials"] = parts[:8]

    ingredients = row.get("ingredients", row.get("active_ingredients", ""))
    if ingredients:
        attrs["ingredients"] = str(ingredients)[:2000]
        if re.search(r"ceramic|glass|porcelain", str(ingredients), re.I):
            attrs["food_contact_material"] = "ceramic"

    safety = row.get("safety_warning", "")
    if safety:
        attrs["safety_warning"] = str(safety)[:2000]

    audience = str(row.get("target_audience", "") or "").lower()
    if re.search(r"child|kid|baby|infant|toddler", audience) or re.search(r"child|kid|baby|infant|toddler", title):
        attrs["childrens_product"] = True

    for key in ("package_weight_g", "item_weight_g", "package_length_cm", "package_width_cm", "package_height_cm"):
        v = row.get(key, "")
        if v not in (None, ""):
            try:
                attrs[key] = float(str(v).replace(" g", "").replace(" cm", "").strip())
            except ValueError:
                pass

    for key in ("asin", "upc", "ean", "gtin", "part_number", "brand", "manufacturer", "model"):
        v = row.get(key, "")
        if v not in (None, ""):
            attrs[key] = str(v)

    return attrs


def normalise_row(row: dict) -> dict:
    def first(*keys, default=""):
        for k in keys:
            v = row.get(k)
            if v not in (None, ""):
                return v
        return default

    sku = first("sku", "asin", "asins", "product_id", "product_id_value", "id", "contribution_sku", default="")
    name = first("name", "title", "item_name", "product_name", default="")
    sku = str(sku).strip()
    name = str(name).strip()
    if not sku or not name:
        raise ValueError("Row missing required sku/name")

    category_str = first("category", "category_root", "root_category", "categories_root", default="")
    item_type = first("item_type", "item_type_keyword", "product_type", "gl_name", default="")
    category = infer_category(category_str, item_type, name)

    market_raw = str(first("market", "marketplace", "locale", default="US"))
    market_map = {"com": "US", "us": "US", "uk": "UK", "de": "EU", "fr": "EU", "it": "EU",
                  "es": "EU", "nl": "EU", "ca": "CA", "au": "AU", "jp": "JP", "mx": "MX"}
    market = market_map.get(market_raw.strip().lower(), market_raw.strip().upper() or "US")

    attrs = infer_attributes(row)
    known = {"sku", "name", "category", "market", "title", "asin", "item_name", "product_id",
             "contribution_sku", "categories_root", "category_root", "root_category",
             "item_type", "item_type_keyword", "product_type", "gl_name", "marketplace", "locale",
             "description", "features", "parent_title", "image", "main_image_link"}
    for k, v in row.items():
        if k in known or v in (None, ""):
            continue
        if isinstance(v, (str, int, float, bool)) and len(str(v)) < 500:
            attrs.setdefault(k, v)

    return {"sku": sku, "name": name, "category": category,
            "market": market, "attributes": attrs}


def _clean_header(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    s = re.sub(r"[^A-Za-z0-9]+", "_", s).strip("_").lower()
    return s


# --------------------------------------------------------------------------
# Excel Parsers (.xlsx, .xlsm, .xlsb)
# --------------------------------------------------------------------------
def _parse_xlsx(path: Path, filename: str) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    fmt = "generic"
    try:
        first_row = _first_row(wb[wb.sheetnames[0]])
        if any(re.search(r"keepa", s, re.I) for s in wb.sheetnames) or "Locale" in first_row:
            fmt = "keepa"
        elif "MAIN_DATA" in wb.sheetnames:
            fmt = "cdq"
        elif "Template" in wb.sheetnames:
            fmt = "amazon_template"
    except Exception:
        fmt = "generic"

    try:
        if fmt == "keepa":
            ws = wb[wb.sheetnames[0]]
            header_row, data_start = 0, 1
        elif fmt == "cdq":
            ws = wb["MAIN_DATA"]
            header_row, data_start = 8, 9
        elif fmt == "amazon_template":
            ws = wb["Template"]
            header_row, data_start = 3, 5
        else:
            ws = wb[wb.sheetnames[0]]
            header_row, data_start = 0, 1

        it = ws.iter_rows(values_only=True)
        header = None
        rows_out: list[dict] = []
        for i, row in enumerate(it):
            if i == header_row:
                header = [_clean_header(v) for v in row]
                continue
            if i < data_start:
                continue
            if not any(v not in (None, "") for v in row):
                continue
            raw = {}
            for j, h in enumerate(header):
                if not h:
                    continue
                v = row[j] if j < len(row) else None
                if v is not None:
                    raw[h] = v
            try:
                rows_out.append(normalise_row(raw))
            except ValueError:
                continue
        return rows_out
    finally:
        wb.close()


def _first_row(ws):
    try:
        for row in ws.iter_rows(values_only=True):
            return [str(v) if v is not None else "" for v in row]
        return []
    except Exception:
        return []


def _parse_xlsb(path: Path) -> list[dict]:
    import pyxlsb
    with pyxlsb.open_workbook(str(path)) as wb:
        for sheet in wb.sheets:
            with wb.get_sheet(sheet) as ws:
                rows_iter = ws.rows()
                try:
                    first = next(rows_iter)
                except StopIteration:
                    continue
                header = [_clean_header(c.v) for c in first]
                joined = " ".join(header)
                if not any(k in joined for k in ("asin", "sku", "title", "name", "product")):
                    continue
                rows_out = []
                for row in rows_iter:
                    raw = {}
                    for j, h in enumerate(header):
                        if not h:
                            continue
                        v = row[j].v if j < len(row) else None
                        if v is not None:
                            raw[h] = v
                    if not any(v not in (None, "") for v in raw.values()):
                        continue
                    try:
                        rows_out.append(normalise_row(raw))
                    except ValueError:
                        continue
                return rows_out
    return []


# --------------------------------------------------------------------------
# Markdown Table Parser (.md, .markdown)
# --------------------------------------------------------------------------
def _parse_markdown_bytes(data: bytes) -> list[dict]:
    text = data.decode("utf-8-sig", errors="replace")
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    rows_out: list[dict] = []
    i = 0
    while i < len(lines):
        line = lines[i]
        if line.startswith("|") and line.endswith("|") and "|" in line[1:-1]:
            header_cols = [c.strip() for c in line.strip("|").split("|")]
            clean_headers = [_clean_header(c) for c in header_cols]
            i += 1
            if i < len(lines) and re.match(r"^\|?\s*:?-+:?\s*(\|?\s*:?-+:?\s*)*\|?$", lines[i]):
                i += 1
            while i < len(lines) and lines[i].startswith("|") and lines[i].endswith("|"):
                row_cols = [c.strip() for c in lines[i].strip("|").split("|")]
                raw = {}
                for j, h in enumerate(clean_headers):
                    if h and j < len(row_cols):
                        raw[h] = row_cols[j]
                if any(raw.values()):
                    try:
                        rows_out.append(normalise_row(raw))
                    except ValueError:
                        pass
                i += 1
            continue
        i += 1
    return rows_out


# --------------------------------------------------------------------------
# PDF & Word Document Parsers (.pdf, .docx)
# --------------------------------------------------------------------------
def _parse_pdf_bytes(path: Path) -> list[dict]:
    rows_out: list[dict] = []
    try:
        import pdfplumber
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables() or []
                for table in tables:
                    if not table or len(table) < 2:
                        continue
                    header = [_clean_header(c or "") for c in table[0]]
                    for row in table[1:]:
                        if not any(row):
                            continue
                        raw = {}
                        for j, h in enumerate(header):
                            if h and j < len(row):
                                raw[h] = str(row[j] or "").strip()
                        if any(raw.values()):
                            try:
                                rows_out.append(normalise_row(raw))
                            except ValueError:
                                pass
    except Exception:
        pass
    return rows_out


def _parse_docx_bytes(path: Path) -> list[dict]:
    rows_out: list[dict] = []
    try:
        import docx
        doc = docx.Document(path)
        for table in doc.tables:
            if not table.rows or len(table.rows) < 2:
                continue
            header = [_clean_header(cell.text or "") for cell in table.rows[0].cells]
            for row in table.rows[1:]:
                cells = [c.text.strip() for c in row.cells]
                if not any(cells):
                    continue
                raw = {}
                for j, h in enumerate(header):
                    if h and j < len(cells):
                        raw[h] = cells[j]
                if any(raw.values()):
                    try:
                        rows_out.append(normalise_row(raw))
                    except ValueError:
                        pass
    except Exception:
        pass
    return rows_out


# --------------------------------------------------------------------------
# Delimited Text (CSV / TSV / TAB / TXT) & JSON
# --------------------------------------------------------------------------
def _parse_delimited_bytes(data: bytes, default_delim: str = ",") -> list[dict]:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = data.decode("utf-8", errors="replace")

    lines = [l for l in text.splitlines() if l.strip() and not l.strip().startswith(("#", "//", "!"))]
    if not lines:
        return []

    sample = "\n".join(lines[:10])
    delim = default_delim
    if "\t" in sample and sample.count("\t") > sample.count(","):
        delim = "\t"
    elif "|" in sample and sample.count("|") > sample.count(","):
        delim = "|"
    elif ";" in sample and sample.count(";") > sample.count(","):
        delim = ";"

    reader = csv.DictReader(io.StringIO("\n".join(lines)), delimiter=delim)
    rows = []
    for r in reader:
        if r and any((v or "").strip() for v in r.values()):
            raw = {_clean_header(k): (v or "").strip() for k, v in r.items() if k}
            try:
                rows.append(normalise_row(raw))
            except ValueError:
                continue
    return rows


def _parse_json_bytes(data: bytes) -> list[dict]:
    text = data.decode("utf-8-sig", errors="replace")
    obj = json.loads(text)
    if isinstance(obj, dict):
        if "products" in obj and isinstance(obj["products"], list):
            obj = obj["products"]
        elif "items" in obj and isinstance(obj["items"], list):
            obj = obj["items"]
        elif "data" in obj and isinstance(obj["data"], list):
            obj = obj["data"]
        else:
            obj = [obj]
    rows = []
    for r in obj:
        if not isinstance(r, dict):
            continue
        raw = {_clean_header(k): v for k, v in r.items()}
        try:
            rows.append(normalise_row(raw))
        except ValueError:
            continue
    return rows


def parse_catalog(path: Path, filename: str) -> list[dict]:
    ext = Path(filename).suffix.lower()
    base = Path(filename).stem.lower()
    if "base" == base or base.startswith("base.") or "asanadata" in base or "asana" in base and ext in (".xlsb", ".xlsx"):
        return []

    if ext in (".xlsx", ".xlsm"):
        return _parse_xlsx(path, filename)
    if ext == ".xlsb":
        return _parse_xlsb(path)
    if ext in (".md", ".markdown"):
        return _parse_markdown_bytes(path.read_bytes())
    if ext == ".docx":
        return _parse_docx_bytes(path)
    if ext == ".pdf":
        return _parse_pdf_bytes(path)
    if ext == ".json":
        return _parse_json_bytes(path.read_bytes())
    if ext in (".ndjson", ".jsonl"):
        rows = []
        for line in path.read_bytes().decode("utf-8-sig", errors="replace").splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                raw = {_clean_header(k): v for k, v in json.loads(line).items() if isinstance(v, (str, int, float, bool))}
                rows.append(normalise_row(raw))
            except (ValueError, json.JSONDecodeError):
                continue
        return rows

    delim = "\t" if ext in (".tsv", ".tab") else ","
    return _parse_delimited_bytes(path.read_bytes(), delim)
