"""Suggested-vs-live listing content comparison.

Local-first sources are ingested into normalized suggested content records. Live
listing snapshots are provider-scoped, timestamped, and only considered fresh
for 48 hours. Comparisons never publish or mutate marketplace content.
"""
from __future__ import annotations

import csv
import difflib
import hashlib
import io
import json
import re
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from fastapi import APIRouter, File, HTTPException, UploadFile

import storage

router = APIRouter(prefix="/api/listings", tags=["listing-comparison"])
FRESHNESS_SECONDS = 48 * 60 * 60
CONTENT_FIELDS = ("title", "item_highlights", "feature_bullet_1", "feature_bullet_2", "feature_bullet_3", "feature_bullet_4", "feature_bullet_5", "description")
ALIASES = {
    "asin": ("asin", "asins", "external_product_id"),
    "sku": ("sku", "item_sku", "seller_sku"),
    "marketplace": ("marketplace", "locale", "market"),
    "title": ("title", "item_name", "product_name", "keepa_title"),
    "item_highlights": ("item_highlights", "highlights"),
    "feature_bullet_1": ("feature_bullet_1", "bullet_point1", "bp1", "extract_bp1", "keepa_bp1"),
    "feature_bullet_2": ("feature_bullet_2", "bullet_point2", "bp2", "extract_bp2", "keepa_bp2"),
    "feature_bullet_3": ("feature_bullet_3", "bullet_point3", "bp3", "extract_bp3", "keepa_bp3"),
    "feature_bullet_4": ("feature_bullet_4", "bullet_point4", "bp4", "extract_bp4", "keepa_bp4"),
    "feature_bullet_5": ("feature_bullet_5", "bullet_point5", "bp5", "extract_bp5", "keepa_bp5"),
    "description": ("description", "product_description", "extract_description", "keepa_description"),
}


def init_listing_compare_db() -> None:
    conn = storage._conn()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS listing_content_sources (
          id TEXT PRIMARY KEY, label TEXT NOT NULL, source_kind TEXT NOT NULL,
          status TEXT NOT NULL DEFAULT 'ready', config TEXT NOT NULL DEFAULT '{}',
          last_success_at TEXT DEFAULT '', last_error TEXT DEFAULT '', created_at TEXT NOT NULL, updated_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS suggested_listing_content (
          id INTEGER PRIMARY KEY AUTOINCREMENT, source_id TEXT NOT NULL, asin TEXT DEFAULT '', sku TEXT DEFAULT '', marketplace TEXT DEFAULT 'US',
          content TEXT NOT NULL, source_observed_at TEXT DEFAULT '', source_locator TEXT DEFAULT '', content_hash TEXT NOT NULL,
          created_at TEXT NOT NULL, updated_at TEXT NOT NULL, UNIQUE(source_id, asin, sku, marketplace)
        );
        CREATE INDEX IF NOT EXISTS idx_suggested_listing_lookup ON suggested_listing_content(asin, sku, marketplace);
        CREATE TABLE IF NOT EXISTS live_listing_snapshots (
          id INTEGER PRIMARY KEY AUTOINCREMENT, source_id TEXT NOT NULL, source_kind TEXT NOT NULL, authority TEXT NOT NULL,
          asin TEXT DEFAULT '', sku TEXT DEFAULT '', marketplace TEXT DEFAULT 'US', seller_id TEXT DEFAULT '',
          source_observed_at TEXT DEFAULT '', retrieved_at TEXT NOT NULL, fresh_until TEXT NOT NULL, status TEXT NOT NULL,
          content TEXT NOT NULL DEFAULT '{}', content_hash TEXT NOT NULL DEFAULT '', request_id TEXT DEFAULT '', error TEXT DEFAULT '', created_at TEXT NOT NULL,
          UNIQUE(source_id, asin, sku, marketplace, retrieved_at)
        );
        CREATE INDEX IF NOT EXISTS idx_live_listing_freshness ON live_listing_snapshots(asin, sku, marketplace, fresh_until);
        CREATE TABLE IF NOT EXISTS listing_content_comparisons (
          id INTEGER PRIMARY KEY AUTOINCREMENT, suggested_id INTEGER NOT NULL, live_snapshot_id INTEGER,
          asin TEXT DEFAULT '', sku TEXT DEFAULT '', marketplace TEXT DEFAULT 'US', field_name TEXT NOT NULL,
          suggested_value TEXT DEFAULT '', live_value TEXT DEFAULT '', exact_match INTEGER NOT NULL DEFAULT 0,
          levenshtein_similarity REAL, phonetic_match INTEGER NOT NULL DEFAULT 0, fuzzy_similarity REAL,
          match_status TEXT NOT NULL, recommendation TEXT NOT NULL DEFAULT '', comparison_status TEXT NOT NULL DEFAULT 'complete',
          created_at TEXT NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_listing_comparison_lookup ON listing_content_comparisons(asin, sku, marketplace, created_at);
    """)
    conn.commit()


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _iso(dt: datetime | None = None) -> str:
    return (dt or _now()).isoformat().replace("+00:00", "Z")


def _parse_iso(value: str | None) -> datetime | None:
    try:
        text = str(value or "").strip()
        if not text: return None
        dt = datetime.fromisoformat(text.replace("Z", "+00:00"))
        return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
    except ValueError:
        return None


def _clean_key(key: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(key or "").lower()).strip("_")


def _first(row: dict, aliases: tuple[str, ...]) -> str:
    for alias in aliases:
        v = row.get(alias)
        if v not in (None, ""):
            return str(v).strip()
    return ""


def _normalized_record(row: dict) -> dict:
    clean = {_clean_key(k): v for k, v in row.items()}
    content = {key: _first(clean, aliases) for key, aliases in ALIASES.items() if key not in {"asin", "sku", "marketplace"}}
    return {
        "asin": _first(clean, ALIASES["asin"]).upper(),
        "sku": _first(clean, ALIASES["sku"]),
        "marketplace": (_first(clean, ALIASES["marketplace"]) or "US").upper(),
        "content": {k: v for k, v in content.items() if v},
        "source_observed_at": str(clean.get("loaded_at") or clean.get("source_observed_at") or ""),
    }


def _parse_upload(raw: bytes, filename: str) -> list[dict]:
    ext = Path(filename).suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        import openpyxl
        from tempfile import NamedTemporaryFile
        with NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            tmp.write(raw); tmp_path = Path(tmp.name)
        try:
            wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
            best = None
            suggested_sheet_priority = {
                "suggested refreshable": 1_000_000,
                "connected-sugg-content": 900_000,
                "suggested export": 800_000,
            }
            # Workbooks often include a live export and a derived comparison
            # tab. For a suggested-content upload, select only a real suggested
            # source table, preferring the workbook's refreshable source.
            for ws in wb.worksheets:
                normalized_sheet = ws.title.strip().lower()
                priority = suggested_sheet_priority.get(normalized_sheet, 0)
                if priority == 0:
                    continue
                sample = list(ws.iter_rows(min_row=1, max_row=20, values_only=True))
                for header_idx, header_row in enumerate(sample):
                    headers = [_clean_key(v) for v in header_row]
                    has_id = any(h in {"asin", "asins", "sku", "item_sku", "external_product_id"} for h in headers)
                    content_aliases = {alias for vals in ALIASES.values() for alias in vals}
                    content_score = sum(h in content_aliases for h in headers)
                    if not has_id or content_score < 2:
                        continue
                    nonempty_content = 0
                    for values in sample[header_idx + 1:]:
                        if any(values[i] not in (None, "") for i, h in enumerate(headers) if h in content_aliases and i < len(values)):
                            nonempty_content += 1
                    score = priority + content_score * 100 + nonempty_content - header_idx
                    if best is None or score > best[0]:
                        best = (score, ws.title, header_idx + 1, headers)
            if not best:
                return []
            _, sheet_name, header_row_num, headers = best
            ws = wb[sheet_name]
            rows = []
            for values in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
                if any(v not in (None, "") for v in values):
                    rows.append({headers[i]: values[i] if i < len(values) else "" for i in range(len(headers)) if headers[i]})
            wb.close()
            return rows
        finally:
            tmp_path.unlink(missing_ok=True)
    if ext in (".json", ".jsonl", ".ndjson"):
        text = raw.decode("utf-8-sig", errors="replace")
        if ext == ".json":
            obj = json.loads(text)
            if isinstance(obj, dict): obj = obj.get("products") or obj.get("items") or obj.get("data") or [obj]
            return [x for x in obj if isinstance(x, dict)]
        return [json.loads(line) for line in text.splitlines() if line.strip()]
    text = raw.decode("utf-8-sig", errors="replace")
    delimiter = "\t" if ext in (".tsv", ".tab") or text.count("\t") > text.count(",") else ","
    return list(csv.DictReader(io.StringIO(text), delimiter=delimiter))


def _levenshtein_similarity(left: str, right: str) -> float:
    if left == right: return 1.0
    if not left or not right: return 0.0
    previous = list(range(len(right) + 1))
    for i, a in enumerate(left, 1):
        current = [i]
        for j, b in enumerate(right, 1):
            current.append(min(current[-1] + 1, previous[j] + 1, previous[j - 1] + (a != b)))
        previous = current
    return 1.0 - previous[-1] / max(len(left), len(right))


def _soundex(text: str) -> str:
    letters = re.sub(r"[^A-Za-z]", "", text).upper()
    if not letters: return ""
    table = {**{c: "1" for c in "BFPV"}, **{c: "2" for c in "CGJKQSXZ"}, **{c: "3" for c in "DT"}, **{c: "4" for c in "L"}, **{c: "5" for c in "MN"}, **{c: "6" for c in "R"}}
    out, previous = [letters[0]], table.get(letters[0], "")
    for c in letters[1:]:
        code = table.get(c, "")
        if code and code != previous: out.append(code)
        previous = code
    return ("".join(out) + "000")[:4]


def _compare_value(suggested: str, live: str) -> dict:
    a = re.sub(r"\s+", " ", suggested.lower()).strip()
    b = re.sub(r"\s+", " ", live.lower()).strip()
    exact = bool(a and a == b)
    lev = _levenshtein_similarity(a, b)
    fuzzy = difflib.SequenceMatcher(None, a, b).ratio()
    phonetic = bool(a and b and _soundex(a) == _soundex(b))
    score = max(lev, fuzzy, 0.9 if phonetic else 0.0)
    if exact: status = "match"
    elif not live: status = "missing_live"
    elif not suggested: status = "missing_suggested"
    elif score >= 0.9: status = "near_match"
    else: status = "mismatch"
    recommendation = "" if status == "match" else (
        "Refresh the live source before applying content." if status == "missing_live" else
        "Review the proposed value against the fresh live listing; keep the stronger compliant copy." if status in {"near_match", "mismatch"} else
        "Add a suggested value before comparison."
    )
    return {"exact_match": exact, "levenshtein_similarity": round(lev, 4), "fuzzy_similarity": round(fuzzy, 4), "phonetic_match": phonetic, "match_status": status, "recommendation": recommendation}


def _latest_live(asin: str, sku: str, marketplace: str) -> dict | None:
    row = storage._conn().execute(
        """SELECT * FROM live_listing_snapshots WHERE marketplace=? AND status='success'
           AND ((? != '' AND asin=?) OR (? != '' AND sku=?)) ORDER BY retrieved_at DESC LIMIT 1""",
        (marketplace, asin, asin, sku, sku),
    ).fetchone()
    if not row: return None
    item = dict(row); item["content"] = json.loads(item["content"] or "{}")
    return item


def _store_snapshot(source_id: str, source_kind: str, authority: str, record: dict, status: str = "success", error: str = "") -> int:
    now = _now(); content = record.get("content") or {}
    digest = hashlib.sha256(json.dumps(content, sort_keys=True).encode()).hexdigest()
    cur = storage._conn().execute(
        """INSERT INTO live_listing_snapshots (source_id,source_kind,authority,asin,sku,marketplace,seller_id,source_observed_at,retrieved_at,fresh_until,status,content,content_hash,error,created_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (source_id, source_kind, authority, record.get("asin", ""), record.get("sku", ""), record.get("marketplace", "US"), "", record.get("source_observed_at", ""), _iso(now), _iso(now + timedelta(seconds=FRESHNESS_SECONDS)), status, json.dumps(content), digest, error, _iso(now)),
    )
    storage._conn().commit(); return cur.lastrowid


@router.post("/suggested/upload")
async def upload_suggested(file: UploadFile = File(...)) -> dict:
    init_listing_compare_db()
    raw = await file.read()
    if len(raw) > 50 * 1024 * 1024: raise HTTPException(413, "Suggested content upload exceeds 50MB")
    rows = _parse_upload(raw, file.filename or "suggested.csv")
    source_id = uuid.uuid4().hex
    now = storage.now_iso()
    conn = storage._conn()
    conn.execute("INSERT INTO listing_content_sources VALUES (?,?,?,?,?,?,?,?,?)", (source_id, Path(file.filename or "suggested").stem, "upload", "ready", "{}", now, "", now, now))
    accepted = 0
    for row in rows:
        record = _normalized_record(row)
        if not (record["asin"] or record["sku"]) or not record["content"]: continue
        digest = hashlib.sha256(json.dumps(record["content"], sort_keys=True).encode()).hexdigest()
        conn.execute("INSERT INTO suggested_listing_content (source_id,asin,sku,marketplace,content,source_observed_at,source_locator,content_hash,created_at,updated_at) VALUES (?,?,?,?,?,?,?,?,?,?)", (source_id, record["asin"], record["sku"], record["marketplace"], json.dumps(record["content"]), record["source_observed_at"], file.filename or "", digest, now, now))
        accepted += 1
    conn.execute("UPDATE listing_content_sources SET status='ready',last_success_at=?,updated_at=? WHERE id=?", (now, now, source_id)); conn.commit()
    return {"source_id": source_id, "rows_read": len(rows), "records_accepted": accepted}


@router.post("/live/upsert")
def upsert_live(body: dict) -> dict:
    """Ingest a normalized live snapshot from an approved adapter or test source."""
    init_listing_compare_db()
    record = _normalized_record(body.get("record") or body)
    if not (record["asin"] or record["sku"]): raise HTTPException(400, "ASIN or SKU is required")
    source_id = str(body.get("source_id") or "manual-live")
    snapshot_id = _store_snapshot(source_id, str(body.get("source_kind") or "manual"), str(body.get("authority") or "user_asserted"), record)
    return {"ok": True, "snapshot_id": snapshot_id, "fresh_until": _latest_live(record["asin"], record["sku"], record["marketplace"])["fresh_until"]}


@router.post("/refresh")
def refresh(body: dict) -> dict:
    """Refresh stale Suggested-vs-Live targets. Keepa is the current live adapter."""
    init_listing_compare_db()
    source_id = body.get("source_id")
    conn = storage._conn()
    sql, params = "SELECT * FROM suggested_listing_content", []
    if source_id: sql += " WHERE source_id=?"; params.append(source_id)
    suggested = [dict(r) for r in conn.execute(sql, params).fetchall()]
    stale = []
    now = _now()
    for row in suggested:
        live = _latest_live(row["asin"], row["sku"], row["marketplace"])
        fresh = live and _parse_iso(live["fresh_until"]) and _parse_iso(live["fresh_until"]) > now
        if not fresh: stale.append(row)
    if not stale: return {"requested": len(suggested), "stale": 0, "refreshed": 0, "status": "fresh"}
    try:
        import keepa
        cfg = keepa._load_config()
        if not cfg.get("api_key"): raise HTTPException(409, "Keepa is not configured; no live provider is available for stale ASINs.")
        asins = list(dict.fromkeys(r["asin"] for r in stale if r["asin"]))
        refreshed = 0
        for i in range(0, len(asins), 100):
            result = keepa.lookup({"asins": ",".join(asins[i:i + 100]), "domain": cfg["domain"], "refresh": True})
            for product in result.get("products", []):
                content = {"title": product.get("title", ""), "item_highlights": "\n".join(product.get("features") or []), **{f"feature_bullet_{n + 1}": v for n, v in enumerate((product.get("features") or [])[:5])}}
                _store_snapshot("keepa", "keepa", "market_intelligence", {"asin": product.get("asin", ""), "marketplace": product.get("domain", "US"), "content": content, "source_observed_at": str(product.get("lastUpdate") or "")})
                refreshed += 1
        return {"requested": len(suggested), "stale": len(stale), "refreshed": refreshed, "status": "success", "authority": "market_intelligence"}
    except HTTPException:
        raise


@router.post("/compare")
def compare(body: dict) -> dict:
    init_listing_compare_db()
    source_id = body.get("source_id")
    strict_fresh = bool(body.get("strict_fresh", True))
    conn = storage._conn()
    sql, params = "SELECT * FROM suggested_listing_content", []
    if source_id: sql += " WHERE source_id=?"; params.append(source_id)
    suggested_rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    now = _now(); output = []; stale = 0
    for suggested in suggested_rows:
        content = json.loads(suggested["content"] or "{}")
        live = _latest_live(suggested["asin"], suggested["sku"], suggested["marketplace"])
        live_is_fresh = bool(live and _parse_iso(live["fresh_until"]) and _parse_iso(live["fresh_until"]) > now)
        if strict_fresh and not live_is_fresh:
            stale += 1; continue
        live_content = (live or {}).get("content") or {}
        for field, value in content.items():
            if field not in CONTENT_FIELDS: continue
            scores = _compare_value(str(value), str(live_content.get(field) or ""))
            status = "stale" if not live_is_fresh else scores["match_status"]
            comparison_status = "complete" if live_is_fresh else "stale"
            rec = "Refresh live data before comparing." if status == "stale" else scores["recommendation"]
            cur = conn.execute("INSERT INTO listing_content_comparisons (suggested_id,live_snapshot_id,asin,sku,marketplace,field_name,suggested_value,live_value,exact_match,levenshtein_similarity,phonetic_match,fuzzy_similarity,match_status,recommendation,comparison_status,created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", (suggested["id"], live["id"] if live else None, suggested["asin"], suggested["sku"], suggested["marketplace"], field, str(value), str(live_content.get(field) or ""), int(scores["exact_match"]), scores["levenshtein_similarity"], int(scores["phonetic_match"]), scores["fuzzy_similarity"], status, rec, comparison_status, storage.now_iso()))
            output.append({"id": cur.lastrowid, "asin": suggested["asin"], "sku": suggested["sku"], "marketplace": suggested["marketplace"], "field": field, "suggested": value, "live": live_content.get(field) or "", "fresh": live_is_fresh, **scores, "match_status": status, "recommendation": rec})
    conn.commit()
    return {"count": len(output), "stale_records": stale, "comparison_status": "complete" if not stale else ("partial" if output else "stale"), "rows": output}


@router.get("/sources")
def sources() -> dict:
    init_listing_compare_db()
    return {"sources": [dict(r) for r in storage._conn().execute("SELECT * FROM listing_content_sources ORDER BY updated_at DESC").fetchall()]}


@router.get("/comparison-report")
def comparison_report(limit: int = 200) -> dict:
    init_listing_compare_db()
    rows = [dict(r) for r in storage._conn().execute("SELECT * FROM listing_content_comparisons ORDER BY created_at DESC LIMIT ?", (min(max(limit, 1), 1000),)).fetchall()]
    return {"count": len(rows), "rows": rows}
